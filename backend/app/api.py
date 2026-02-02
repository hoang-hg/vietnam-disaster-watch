import logging
from fastapi import APIRouter, Depends, Query

logger = logging.getLogger(__name__)
from sqlalchemy import desc, func, or_, case
from sqlalchemy.orm import Session, defer, selectinload
from sqlalchemy.sql import text
from .database import get_db, engine
from . import models
from .models import Article, Event, Blacklist, CrawlerStatus, AiFeedback
from .schemas import ArticleOut, EventOut, EventDetailOut, EventUpdate
from datetime import datetime, timedelta, timezone
from fastapi import Response, Request, HTTPException, status
from fastapi.responses import StreamingResponse, HTMLResponse
from .event_matcher import upsert_event_for_article, emit_event_notifications
from .log_utils import log_audit
from .settings import settings
import asyncio
from pathlib import Path
import json
from .nlp import PROVINCES, strip_accents
from .sources import DISASTER_GROUPS
from .risk_lookup import canon
from . import broadcast, auth
from .cache import cache
import io
from typing import Optional, List

# Unified filtering rules for Dashboard/Stats (Decision 18/2021/QĐ-TTg)
def get_dec18_filter():
    """Returns the SQLAlchemy filter condition for Decision 18."""
    major_hazards = list(DISASTER_GROUPS.keys())
    return or_(
        Event.disaster_type.in_(major_hazards),
        (func.coalesce(Event.deaths, 0) + func.coalesce(Event.missing, 0) + 
         func.coalesce(Event.injured, 0) + func.coalesce(Event.damage_billion_vnd, 0)) > 0,
        Event.confidence >= 1.0 # Include Admin/Community verified events 
    )

def apply_dashboard_filters(query, db: Session):
    """Applies Decision 18 filters directly to a SQLAlchemy query."""
    major_hazards = list(DISASTER_GROUPS.keys())
    

    dec18_filter = get_dec18_filter()
    
    # Remove 'community' from strict exclusion so verified community events can pass via dec18_filter
    return query.filter(Event.disaster_type != "unknown")\
                .filter(Event.sources_count > 0)\
                .filter(dec18_filter)

# --- HELPER FUNCTIONS & CONSTANTS ---

TYPE_MAP = {
    "storm": "Bão, ATNĐ",
    "flood": "Lũ lụt",
    "flash_flood": "Lũ quét, Lũ ống",
    "landslide": "Sạt lở",
    "subsidence": "Sụt lún đất",
    "drought": "Hạn hán",
    "salinity": "Xâm nhập mặn",
    "extreme_weather": "Mưa lớn, Lốc, Sét, Mưa Đá",
    "heatwave": "Nắng nóng",
    "cold_surge": "Rét hại, Sương muối",
    "earthquake": "Động đất",
    "tsunami": "Sóng thần",
    "storm_surge": "Nước dâng",
    "wildfire": "Cháy rừng",
    "erosion": "Xói lở",
    "warning_forecast": "Cảnh báo, dự báo",
    "recovery": "Khắc phục hậu quả",
    "community": "Cộng đồng báo cáo",
    "unknown": "Chưa phân loại"
}

def get_date_range(hours: int, date: str | None, start_date: str | None, end_date: str | None):
    """Standardizes date range parsing for all stats endpoints (Timezone Adjusted)."""
    vn_tz = timezone(timedelta(hours=7))
    
    def to_utc_naive(dt_vn_naive: datetime) -> datetime:
        # Treat input naive (from strptime) as VN time, convert to UTC, then strip tz for DB
        return dt_vn_naive.replace(tzinfo=vn_tz).astimezone(timezone.utc).replace(tzinfo=None)

    if start_date or end_date:
        try:
            # [LOGIC CHANGE] Interpret inputs as VN time
            
            if start_date:
                s_vn = datetime.strptime(start_date, "%Y-%m-%d")
                start = to_utc_naive(s_vn)
            else:
                start = datetime.min
            
            # If start_date is provided but end_date is NOT, treat as a single day filter (legacy behavior)
            if start_date and not end_date:
                s_vn = datetime.strptime(start_date, "%Y-%m-%d")
                e_vn = s_vn + timedelta(days=1)
                end = to_utc_naive(e_vn)
            else:
                if end_date:
                    e_vn = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1) # End of that day
                    end = to_utc_naive(e_vn)
                else:
                    end = datetime.now(timezone.utc).replace(tzinfo=None)
            
            return start, end
        except ValueError:
            pass # Fallback to default functions
            
    elif date:
        try:
            d_vn = datetime.strptime(date, "%Y-%m-%d")
            start = to_utc_naive(d_vn)
            end = to_utc_naive(d_vn + timedelta(days=1))
            return start, end
        except ValueError:
            pass
            
    # Default: Use hours (relative to current UTC)
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    since = now - timedelta(hours=hours)
    return since, now

def get_visibility_filter(query, is_admin: bool):
    """Applies standard visibility rules for non-admin users."""
    if is_admin:
        return query
    
    # Public users:
    # - Show if confidence >= 0.8
    # - OR if needs_verification=0 AND sources_count >= 2
    return query.filter(or_(
        Event.confidence >= 0.8,
        (Event.needs_verification.is_(False)) & (Event.sources_count >= 2)
    ))



router = APIRouter(prefix="/api", tags=["api"])

@router.get("/health")
def health():
    return {"ok": True}

@router.get("/articles/latest", response_model=list[ArticleOut])
def latest_articles(
    limit: int = Query(50, ge=1, le=200),
    type: str | None = Query(None),
    province: str | None = Query(None),
    exclude_unknown: bool = Query(False),
    after_id: int | None = Query(None, description="Keyset pagination for better performance"),
    q: str | None = Query(None),
    db: Session = Depends(get_db),
):
    cache_key = f"articles_latest_{limit}_{type}_{province}_{exclude_unknown}_{after_id}_{q}"
    cached = cache.get(cache_key)
    if cached: return cached

    q_latest = db.query(Article).filter(Article.status == "approved")\
        .options(defer(Article.full_text))\
        .order_by(desc(Article.published_at))

    if q:
        q_clean = strip_accents(q).lower()
        q_latest = q_latest.filter(or_(
            Article.title.ilike(f"%{q}%"),
            Article.title.ilike(f"%{q_clean}%")
        ))

    


    if after_id is not None:
        # [FIX] Keyset pagination by ID is unstable when sorting by published_at.
        # We fallback to a safer filtering or simple offset-like behavior if after_id is used.
        # For now, we use it as a secondary sort key to ensure stability.
        q_latest = q_latest.filter(Article.id < after_id)
    if type:
    	q_latest = q_latest.filter(Article.disaster_type == type)
    if province:
    	q_latest = q_latest.filter(Article.province == province)
    if exclude_unknown:
    	q_latest = q_latest.filter(Article.disaster_type != 'unknown')
    
    res = q_latest.limit(limit).all()
    
    # [OPTIMIZATION] Serialize to dicts before caching to avoid detaching ORM objects
    # and to ensure the cache stores simple serializable data.
    # We use jsonable_encoder or model_dump depending on need, but since ArticleOut is a Pydantic model:
    serialized = [ArticleOut.model_validate(a).model_dump() for a in res]
    
    cache.set(cache_key, serialized, ttl=120)
    return res

    # Optimized filtering logic: move simple filters to DB to reduce payload
def get_base_event_query(db: Session):
    return db.query(Event)

@router.get("/events")
def events(
    request: Request,
    response: Response,
    limit: int = Query(50, ge=1, le=1000), # PostgreSQL supports >1000. SQLite limit is 999 variables in IN clause.
    offset: int = Query(0, ge=0),
    after_id: int | None = Query(None, description="Keyset pagination for better performance on large datasets"),
    hours: int | None = Query(None, ge=1, le=720),
    type: str | None = Query(None),
    province: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    q: str | None = Query(None),
    date: str | None = Query(None),
    wrapper: bool = Query(False),
    quick: str | None = Query(None),
    min_lat: float | None = Query(None),
    max_lat: float | None = Query(None),
    min_lon: float | None = Query(None),
    max_lon: float | None = Query(None),
    db: Session = Depends(get_db),
    sort: str = Query("impact"),
    current_user: Optional[models.User] = Depends(auth.get_current_user_optional),
):
    is_admin = current_user and current_user.role == "admin"
    
    # Cache optimization - include is_admin, offset, wrapper and bbox in key
    cache_key = f"ev_v2_{limit}_{offset}_{hours}_{type}_{province}_{start_date}_{end_date}_{q}_{date}_{sort}_{is_admin}_{wrapper}_{quick}_{min_lat}_{max_lat}_{min_lon}_{max_lon}"
    cached = cache.get(cache_key)
    if cached:
        response.headers["X-Cache"] = "HIT"
        # [OPTIMIZATION] Reduce cache time or disable for admins to ensure instant feedback
        if is_admin:
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        else:
            response.headers["Cache-Control"] = "public, max-age=10"
        return cached

    query = db.query(Event)

    # 1. Base Security / Visibility Filter
    query = get_visibility_filter(query, is_admin)

    # 2. Database-level filters (Decision 18/2021/QĐ-TTg Implementation)
    query = apply_dashboard_filters(query, db)

    # 3. BBox Filtering (Map Viewport Optimization)
    if min_lat is not None:
        query = query.filter(Event.lat >= min_lat)
    if max_lat is not None:
        query = query.filter(Event.lat <= max_lat)
    if min_lon is not None:
        query = query.filter(Event.lon >= min_lon)
    if max_lon is not None:
        query = query.filter(Event.lon <= max_lon)

    # Standardized Date Logic
    if date or start_date or end_date:
        # We assume 'hours' is ignored if specific dates are provided, 
        # but the original logic had separate blocks. 
        # For 'events' list specifically, we can reuse get_date_range logic 
        # BUT we must respect the original 'hours' param default of None
        # So we only apply if they are present.
        d_start, d_end = get_date_range(hours or 24, date, start_date, end_date)
        # [OPTIMIZATION] Show events that were active/updated in the window, not just started
        query = query.filter(Event.last_updated_at >= d_start, Event.started_at < d_end)
    elif hours:
        since = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=hours)
        query = query.filter(Event.last_updated_at >= since)

    if type: query = query.filter(Event.disaster_type == type)
    if province: query = query.filter(Event.province == province)
    if q: 
        q_clean = strip_accents(q).lower()
        query = query.filter(or_(
            Event.title.ilike(f"%{q}%"),
            Event.title.ilike(f"%{q_clean}%")
        ))

    # Quick Filters (Decision & UI Shortcuts)
    if quick == "casualties":
        query = query.filter((func.coalesce(Event.deaths, 0) + func.coalesce(Event.missing, 0) + func.coalesce(Event.injured, 0)) > 0)
    elif quick == "damage":
        query = query.filter(func.coalesce(Event.damage_billion_vnd, 0) > 0)
    elif quick == "provinces":
        query = query.filter(Event.province.in_(PROVINCES))
    elif quick == "community":
        query = query.filter(Event.disaster_type == "community")

    # [BBOX FILTER] Map optimization
    if min_lat is not None: query = query.filter(Event.lat >= min_lat)
    if max_lat is not None: query = query.filter(Event.lat <= max_lat)
    if min_lon is not None: query = query.filter(Event.lon >= min_lon)
    if max_lon is not None: query = query.filter(Event.lon <= max_lon)

    # [FIX] Keyset pagination (after_id) is only stable when sorting by a unique, strictly monotonic field (like ID).
    # If sorting by 'impact' or 'latest' (dates can be identical), after_id can cause skipped items or duplicates.
    # We restrict after_id usage to avoid logic errors in complex sorts.
    if after_id is not None and sort not in ["impact", "latest"]:
        query = query.filter(Event.id < after_id)

    # Calculate total if wrapper is requested
    total_count = 0
    if wrapper:
        total_count = query.count()

    # 4. Apply Sorting (MUST be done before Limit/Offset for correct pagination)
    if sort == "latest":
        query = query.order_by(desc(Event.started_at), desc(Event.last_updated_at))
    else:
        # 4. [OPTIMIZATION] Efficient Fetching
        if sort == "start":
            query = query.order_by(desc(Event.started_at))
        else:
            # Sort by impact + confidence
            query = query.order_by(
                desc(func.coalesce(Event.deaths, 0) * 10 + func.coalesce(Event.missing, 0) * 8 + func.coalesce(Event.damage_billion_vnd, 0)),
                desc(Event.confidence)
            )

    filtered = query.limit(limit).offset(offset).all()
    if not filtered:
        if wrapper:
            return {"items": [], "total": total_count}
        return []

    # 5. [OPTIMIZATION] Fix N+1: Batch Fetch Article Counts
    event_ids = [e.id for e in filtered]
    
    # Count approved articles per event in one query
    count_q = db.query(Article.event_id, func.count(Article.id)).filter(
        Article.event_id.in_(event_ids),
        Article.status == "approved"
    )
    if hours:
        h_start = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(hours=hours)
        count_q = count_q.filter(Article.published_at >= h_start)
    
    counts_map = {row[0]: row[1] for row in count_q.group_by(Article.event_id).all()}

    # 6. [OPTIMIZATION] Fix N+1: Batch Fetch Images & Sources
    # Prioritizing 'approved' status then latest publication date
    subq = db.query(
        Article.event_id,
        Article.image_url,
        Article.source,
        Article.url,
        func.row_number().over(
            partition_by=Article.event_id,
            order_by=[
                desc(Article.image_url.isnot(None)), # Prioritize articles WITH images
                desc(Article.status == "approved"), 
                desc(Article.published_at)
            ]
        ).label("rn")
    ).filter(
        Article.event_id.in_(event_ids),
        Article.status.in_(["approved", "pending"])
    ).subquery()
    
    # Optimized fetch of latest image/source per event
    leads = db.query(subq).filter(subq.c.rn == 1).all()
    # Build map: event_id -> (image_url, source, url)
    leads_map = {row.event_id: (row.image_url, row.source, row.url) for row in leads}

    # 7. Final response assembly
    events_out = []
    for ev in filtered:
        # We manually map to avoid multiple DB hits from ORM attributes
        ev_data = EventOut.model_validate(ev)
        ev_id = ev.id
        
        ev_data.articles_count = counts_map.get(ev_id, 0)
        
        img, src, url = leads_map.get(ev_id, (None, None, None))
        ev_data.image_url = img
        ev_data.source = src
        ev_data.source_url = url
        
        # Inject Fallback Image
        if not ev_data.image_url:
            chosen_img = DEFAULT_IMAGES.get(ev.disaster_type, DEFAULT_IMAGES["unknown"])
            if ev.disaster_type == "extreme_weather" and "mưa đá" in (ev.title or "").lower():
                chosen_img = SUB_IMAGES["hail"]
            ev_data.image_url = chosen_img
        
        # [NEW] Check logic: if it has manual location_description, prefer it over province text?
        # Actually EventOut already includes location_description, frontend handles display.

        events_out.append(ev_data)

    # [OPTIMIZATION] Serialize detailed Pydantic models to dicts before caching
    final_result = [e.model_dump() for e in events_out]
    
    if wrapper:
        result = {"items": final_result, "total": total_count}
        cache.set(cache_key, result, ttl=60) # Reduced from 300
        if is_admin:
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        else:
            response.headers["Cache-Control"] = "public, max-age=10"
        return result
    else:
        cache.set(cache_key, final_result, ttl=60) # Reduced from 300
        if is_admin:
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        else:
            response.headers["Cache-Control"] = "public, max-age=10"
        return final_result

@router.get("/share/events/{event_id}", response_class=HTMLResponse)
def share_event_page(event_id: int, db: Session = Depends(get_db)):
    """A minimal HTML page with OpenGraph tags for social crawlers."""
    ev = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not ev:
        raise HTTPException(status_code=404, detail="Không tìm thấy sự kiện")
    
    # Get top article for image context
    top_art = db.query(models.Article).filter(
        models.Article.event_id == event_id, 
        models.Article.image_url.isnot(None), 
        models.Article.status == "approved"
    ).order_by(models.Article.published_at.desc()).first()
    
    # Use event image or article image or default
    raw_img = ev.image_url or (top_art.image_url if top_art else "")
    
    # Ensure absolute URL for social bots
    if raw_img and raw_img.startswith("http"):
        image_url = raw_img
    elif raw_img and raw_img.startswith("/"):
        image_url = f"{settings.frontend_url}{raw_img}"
    else:
        # Generic fallback
        image_url = "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/alert-triangle.svg"
    
    title = f"{ev.title} | Theo dõi thiên tai VDW"
    desc = f"Cảnh báo rủi ro thiên tai tại {ev.province}. Loại: {ev.disaster_type}. Cập nhật tóm tắt thiệt hại và diễn biến mới nhất."
    app_url = f"{settings.frontend_url}/events/{event_id}"

    return f"""
    <!DOCTYPE html>
    <html lang="vi">
    <head>
        <meta charset="UTF-8">
        <title>{title}</title>
        <meta name="description" content="{desc}">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        
        <!-- Open Graph / Facebook / Zalo -->
        <meta property="og:type" content="article">
        <meta property="og:url" content="{app_url}">
        <meta property="og:title" content="{title}">
        <meta property="og:description" content="{desc}">
        <meta property="og:image" content="{image_url}">
        <meta property="og:site_name" content="Viet Disaster Watch">

        <!-- Redirect real users to the React App -->
        <script>
            window.location.href = "{app_url}";
        </script>
        
        <style>
            body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; display: flex; align-items: center; justify-content: center; height: 100vh; margin: 0; background: #f8fafc; color: #1e293b; }}
            .container {{ text-align: center; padding: 2rem; max-width: 400px; }}
            .loader {{ border: 3px solid #e2e8f0; border-top: 3px solid #2fa1b3; border-radius: 50%; width: 32px; height: 32px; animation: spin 0.8s linear infinite; margin: 0 auto 1.5rem; }}
            @keyframes spin {{ 0% {{ transform: rotate(0deg); }} 100% {{ transform: rotate(360deg); }} }}
            a {{ color: #2fa1b3; text-decoration: none; font-weight: 600; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="loader"></div>
            <h2 style="font-size: 1.125rem; font-weight: 700; margin-bottom: 0.5rem;">Đang kết nối...</h2>
            <p style="font-size: 0.875rem; color: #64748b;">Hệ thống đang chuyển hướng bạn đến báo cáo chi tiết về rủi ro thiên tai.</p>
            <p style="font-size: 0.75rem; margin-top: 2rem; color: #94a3b8;">
                Nếu trang không tự chuyển, <a href="{app_url}">nhấn vào đây</a>.
            </p>
        </div>
    </body>
    </html>
    """

# Lightweight SVGs (stable CDN, pinned version)
DEFAULT_IMAGES = {
    "storm": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/cloud-storm.svg",
    "flood": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/droplet.svg",
    "flash_flood": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/droplet.svg",
    "landslide": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/triangle.svg",
    "subsidence": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/arrow-down-circle.svg",
    "drought": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/sun-off.svg",
    "salinity": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/ripple.svg",
    "extreme_weather": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/cloud-lightning.svg",
    "heatwave": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/sun.svg",
    "cold_surge": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/snowflake.svg",
    "earthquake": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/activity.svg",
    "tsunami": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/droplet.svg",
    "storm_surge": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/ripple.svg",
    "wildfire": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/flame.svg",
    "erosion": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/layers-difference.svg",
    "warning_forecast": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/alert-circle.svg",
    "recovery": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/tool.svg",
    "unknown": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/urgent.svg",
}

SUB_IMAGES = {
    # “landslide”: không có icon chuyên “sạt lở” trong bộ v1.13.0, dùng biểu tượng “sườn dốc” (triangle) làm proxy.
    "landslide": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/triangle.svg",

    # “tsunami/storm surge”: bộ v1.13.0 không có “wave” rõ ràng, dùng droplet (nước dâng) làm proxy.
    "tsunami": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/droplet.svg",

    # “hail”: dùng snowflake làm proxy (mưa đá/hạt băng).
    "hail": "https://cdnjs.cloudflare.com/ajax/libs/tabler-icons/1.13.0/icons/snowflake.svg",
}

@router.get("/events/{event_id}", response_model=EventDetailOut)
def event_detail(
    event_id: int, 
    response: Response,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(auth.get_current_user_optional)
):
    """Chi tiết sự kiện + danh sách bài báo (có phân trang nhẹ)."""
    limit_articles = limit
    is_admin = current_user and current_user.role == "admin"
    
    # 1. Try Cache
    cache_key = f"ev_detail_v3_{event_id}_{is_admin}"
    cached = cache.get(cache_key)
    if cached:
        response.headers["X-Cache"] = "HIT"
        response.headers["Cache-Control"] = "public, max-age=60"
        return cached
    
    # 2. Optimized Fetch (Get Event only first)
    query = db.query(Event).filter(Event.id == event_id)
    query = get_visibility_filter(query, is_admin)
    ev = query.first()
    
    if not ev:
        raise HTTPException(status_code=404, detail="Sự kiện không tồn tại hoặc bạn không có quyền xem.")
        
    # 3. [OPTIMIZATION] Database-side Article Filtering & Limited Fetch
    # For large datasets, we should not load thousands of articles into a single response.
    # We fetch the latest 300 articles (status approved/pending).
    # [FIX] Privacy & Verification: Regular users should ONLY see approved articles.
    # Pending articles might contain unverified rumors.
    is_admin = current_user and current_user.role == "admin"
    allowed_statuses = ["approved", "pending"] if is_admin else ["approved"]
    
    articles_q = db.query(Article).filter(
        Article.event_id == event_id,
        Article.status.in_(allowed_statuses)
    ).options(defer(Article.full_text)).order_by(desc(Article.published_at))
    
    # Accurate counts via SQL aggregates
    total_articles = articles_q.count()
    sources_count_val = db.query(func.count(func.distinct(Article.source))).filter(
        Article.event_id == event_id,
        Article.status.in_(["approved", "pending"])
    ).scalar() or 0
    
    # Fetch limited set
    articles = articles_q.limit(limit_articles).all()
    
    # 4. Map to Schema
    ev_data = EventDetailOut.model_validate(ev)
    ev_data.articles = [ArticleOut.model_validate(a) for a in articles]
    ev_data.articles_count = total_articles
    ev_data.sources_count = sources_count_val
    
    # 5. Save to Cache
    result = ev_data.model_dump()
    cache.set(cache_key, result, ttl=300)
    
    response.headers["Cache-Control"] = "public, max-age=60"
    return result

@router.put("/events/{event_id}", response_model=EventOut)
def update_event(
    event_id: int, 
    payload: EventUpdate, 
    db: Session = Depends(get_db),
    admin: models.User = Depends(auth.get_current_admin)
):
    """Update event details (admin only)."""
    ev = db.query(Event).filter(Event.id == event_id).with_for_update().first()
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
        
    update_data = payload.model_dump(exclude_unset=True)
    
    # Audit logging for manual correction
    log_audit("manual_correction", admin.id, admin.email, event_id=event_id, details=update_data)

    # Apply changes
    old_started_at = ev.started_at
    for field, value in update_data.items():
        setattr(ev, field, value)
    

    
    # Propagate date change to articles if started_at changed
    if 'started_at' in update_data and update_data['started_at']:
        try:
            # Pydantic likely parsed this into a datetime, but ensure it's compatible
            raw_date = update_data['started_at']
            if isinstance(raw_date, str):
                raw_date = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
            
            # Normalize to Naive UTC for DB consistency
            if raw_date.tzinfo is not None:
                # Convert to UTC first if it has other timezone, then strip
                new_date = raw_date.astimezone(timezone.utc).replace(tzinfo=None)
            else:
                new_date = raw_date

            # Update Event model explicitly to ensure naive format is set
            ev.started_at = new_date
            
            # Update all linked articles published_at to match the new start date
            # This aligns the event timeline
            res = db.query(models.Article).filter(models.Article.event_id == event_id).update({
                models.Article.published_at: new_date
            }, synchronize_session=False)
            logger.info(f"Propagated event date {new_date} to {res} articles for event {event_id}")
            
        except Exception as e:
            logger.error(f"Failed to propagate date change to articles: {e}")

    # [FIX] Propagate type and province changes to articles for consistency
    if 'disaster_type' in update_data and update_data['disaster_type']:
        try:
            db.query(models.Article).filter(models.Article.event_id == event_id).update({
                models.Article.disaster_type: update_data['disaster_type']
            }, synchronize_session=False)
        except Exception as e:
            logger.error(f"Failed to propagate type change: {e}")

    if 'province' in update_data and update_data['province']:
        try:
            db.query(models.Article).filter(models.Article.event_id == event_id).update({
                models.Article.province: update_data['province']
            }, synchronize_session=False)
        except Exception as e:
            logger.error(f"Failed to propagate province change: {e}")

    # [FIX] Also propagate detailed location fields for consistency in Exports/Search
    loc_fields = ['commune', 'village', 'route', 'landmark']
    loc_updates = {f: update_data[f] for f in loc_fields if f in update_data}
    if loc_updates:
        try:
            db.query(models.Article).filter(models.Article.event_id == event_id).update(
                loc_updates, synchronize_session=False
            )
        except Exception as e:
            logger.error(f"Failed to propagate location updates: {e}")

    ev.last_updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    db.commit()
    db.refresh(ev)
    
    # [SYNC] Broadcast update to all clients
    from .event_matcher import emit_event_notifications
    emit_event_notifications(db, ev.id, is_new=False)

    # Invalidate cache
    cache.delete_match(f"ev_detail_v3_{event_id}*")
    cache.delete_match("ev_v2_*")
    cache.delete_match("stats_*")
    cache.delete_match("timeline_*")
    cache.delete_match("articles_latest_*")
    cache.delete_match("map_*")
    cache.delete_match("heatmap_*")
    
    return ev

@router.delete("/events/{event_id}", status_code=204)
def delete_event(
    event_id: int, 
    db: Session = Depends(get_db),
    admin: models.User = Depends(auth.get_current_admin)
):
    """
    Delete an event (admin only).
    Also updates associated articles to 'rejected' status so they don't reappear.
    """
    ev = db.query(Event).filter(Event.id == event_id).first()
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
        
    # [OPTIMIZATION] Batch check and bulk insert to avoid N+1 queries
    articles = db.query(Article).filter(Article.event_id == event_id).all()
    
    hashes = [a.news_hash for a in articles if a.news_hash]
    if hashes:
        existing_hashes = {
            r[0] for r in db.query(Blacklist.news_hash).filter(Blacklist.news_hash.in_(hashes)).all()
        }
        
        new_blacklist_entries = []
        seen_in_batch = set()
        
        for art in articles:
            if art.news_hash and art.news_hash not in existing_hashes and art.news_hash not in seen_in_batch:
                new_blacklist_entries.append(Blacklist(
                    news_hash=art.news_hash,
                    title=art.title,
                    reason=f"Admin deleted parent event: {ev.title}"
                ))
                seen_in_batch.add(art.news_hash)
        
        if new_blacklist_entries:
            db.add_all(new_blacklist_entries)

    # Mark all associated articles (news + shadow community arts) as Rejected/Hidden
    db.query(Article).filter(Article.event_id == event_id).update(
        {"status": "rejected", "event_id": None}, 
        synchronize_session=False
    )
    
    # [FIX] Also unlink from CrowdsourcedReport if any
    db.query(models.CrowdsourcedReport).filter(models.CrowdsourcedReport.event_id == event_id).update(
        {"event_id": None},
        synchronize_session=False
    )
    
    # Delete the event
    title_for_log = ev.title
    _delete_event_internal(db, ev)
    
    log_audit("delete_event", admin.id, admin.email, event_id=event_id, details={"title": title_for_log})
    
    logger.info(f"Deleted event {event_id}: {title_for_log}")
    
    return

@router.delete("/articles/{article_id}", status_code=204)
def delete_article(
    article_id: int, 
    db: Session = Depends(get_db),
    admin: models.User = Depends(auth.get_current_admin)
):
    """
    Delete/Reject an article (admin only).
    """
    art = _reject_article_internal(db, article_id, admin, "Direct deletion via API")
    if not art:
        raise HTTPException(status_code=404, detail="Article not found")
    return

def recalculate_event_metrics(db: Session, event_id: int):
    """
    Helper to recalculate Event metrics (deaths, injured, sources_count)
    after an admin explicitly removes an article.
    Uses efficient SQL aggregation instead of loading objects.
    """
    ev = db.query(Event).filter(Event.id == event_id).first()
    if not ev: return
    
    # SQL Aggregation for metrics
    aggs = db.query(
        func.max(Article.deaths),
        func.max(Article.missing),
        func.max(Article.injured),
        func.max(Article.damage_billion_vnd),
        func.count(func.distinct(Article.source))
    ).filter(
        Article.event_id == event_id,
        Article.status == "approved"
    ).first()
    
    # Update with aggregated max values
    ev.deaths = aggs[0] or 0
    ev.missing = aggs[1] or 0
    ev.injured = aggs[2] or 0
    ev.damage_billion_vnd = aggs[3] or 0.0
    ev.sources_count = aggs[4] or 0
    
    # Keep Event visual & title fresh: if current image/title missing, take from latest approved
    latest_art = db.query(Article).filter(Article.event_id == event_id, Article.status == "approved")\
        .order_by(Article.published_at.desc()).first()
    if latest_art:
        if not ev.image_url: ev.image_url = latest_art.image_url
        if len(ev.title or "") < 10: ev.title = latest_art.title

    # Downgrade confidence if sources dropped explicitly
    if ev.sources_count < 2 and ev.confidence > 0.6:
        ev.confidence = 0.5
            
    # [FIX] Logic Check: If 0 Approved Sources, check if we have Pending ones.
    # If we have Pending, we CANNOT delete the event via _delete_event_internal because it REJECTS all articles.
    # We should just downgrade it to waiting state.
    if ev.sources_count == 0:
        pending_count = db.query(func.count(Article.id)).filter(
            Article.event_id == event_id, 
            Article.status == "pending"
        ).scalar() or 0
        
        if pending_count > 0:
            ev.needs_verification = True
            ev.confidence = 0.3 # Low confidence, waiting for review
            # Do NOT delete
        else:
            # Truly empty (No Approved, No Pending)
            _delete_event_internal(db, ev)
            return

    db.commit()
    db.refresh(ev)
    
    # [SYNC] Broadcast metric changes to all clients
    emit_event_notifications(db, event_id, is_new=False)
    
    # Clear specific detail caches
    cache.delete(f"ev_detail_v3_{event_id}_True")
    cache.delete(f"ev_detail_v3_{event_id}_False")
    cache.delete_match("ev_v2_*")

def _delete_event_internal(db: Session, ev: Event):
    """Internal helper to shared event deletion logic."""
    event_id = ev.id
    title = ev.title
    
    # [FIX] Blacklist articles from deleted events to prevent re-crawling bad data
        # Fetching hashes before deletion to maintain audit trail/blacklist
    bad_articles = db.query(Article).filter(Article.event_id == event_id).all()
    for art in bad_articles:
        if art.news_hash:
            existing = db.query(models.Blacklist).filter(models.Blacklist.news_hash == art.news_hash).first()
            if not existing:
                db.add(models.Blacklist(
                    news_hash=art.news_hash,
                    title=art.title,
                    reason=f"Rejected because parent event '{title}' was deleted"
                ))

    # Batch unlink articles
    db.query(Article).filter(Article.event_id == event_id).update(
        {"status": "rejected", "event_id": None}, 
        synchronize_session=False
    )

    # [FIX] Unlink Crowdsourced Reports
    # Prevents ForeignKeyViolation: update or delete on table "events" violates foreign key constraint
    # Also set status to 'rejected' so it doesn't count in dashboard stats anymore
    db.query(models.CrowdsourcedReport).filter(models.CrowdsourcedReport.event_id == event_id).update(
        {"event_id": None, "status": "rejected"},
        synchronize_session=False
    )
    
    db.delete(ev)
    db.commit()
    
    # [SYNC] Broadcast delete signal & Clear Caches
    from .ws import manager
    manager.broadcast_sync({"type": "EVENT_DELETE", "id": event_id})
    from .cache import cache
    cache.delete_match(f"ev_detail_v3_{event_id}*")
    cache.delete_match("ev_v2_*")
    cache.delete_match("stats_*")
    cache.delete_match("map_*")

    # Comprehensive cache invalidation
    cache.delete_match(f"ev_detail_v3_{event_id}*")
    cache.delete_match("ev_v2_*")
    cache.delete_match("stats_*")
    cache.delete_match("map_*")
    
    logger.info(f"System deleted event {event_id}: {title}")

def _reject_article_internal(db: Session, article_id: int, admin: models.User, reason: str = "Admin manual removal"):
    """Unified logic for removing/rejecting an article."""
    art = db.query(Article).filter(Article.id == article_id).first()
    if not art:
        return None
        
    old_event_id = art.event_id
    art.status = "rejected"
    art.event_id = None
    art.moderated_by = admin.id
    art.moderated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    
    # Blacklist hash
    if art.news_hash:
        existing = db.query(models.Blacklist).filter(models.Blacklist.news_hash == art.news_hash).first()
        if not existing:
            db.add(models.Blacklist(
                news_hash=art.news_hash,
                title=art.title,
                reason=reason
            ))
            
    log_audit("reject_article", admin.id, admin.email, event_id=old_event_id, article_id=article_id, details={"reason": reason})
    db.commit()
    
    # [FIX] Recalculate metrics for the event after article removal
    if old_event_id:
        recalculate_event_metrics(db, old_event_id)
        db.commit()
        
    # Global cache invalidation for general article lists
    cache.delete_match("articles_latest_*")
    cache.delete_match("stats_*")
    return art



@router.get("/stats/summary")
def stats_summary(
    hours: int = Query(24, ge=1, le=720), 
    date: str | None = Query(None), 
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    type: str | None = Query(None),
    province: str | None = Query(None),
    q: str | None = Query(None),
    quick: str | None = Query(None),
    response: Response = None,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(auth.get_current_user_optional),
):
    is_admin = current_user and current_user.role == "admin"
    cache_key = f"stats_{hours}_{date}_{start_date}_{end_date}_{type}_{province}_{q}_{is_admin}_{quick}"
    cached = cache.get(cache_key)
    if cached:
        if response: response.headers["Cache-Control"] = "public, max-age=120"
        return cached

    start, end = get_date_range(hours, date, start_date, end_date)


    # 1. Crowdsourced reports count (approved)
    community_reports_q = db.query(func.count(models.CrowdsourcedReport.id)).filter(
        models.CrowdsourcedReport.status == "approved",
        models.CrowdsourcedReport.event_id.isnot(None) # [FIX] Only count if linked to an active event
    )
    # [FIX] Use standardized time range
    community_reports_q = community_reports_q.filter(
        models.CrowdsourcedReport.created_at >= start,
        models.CrowdsourcedReport.created_at < end
    )
    
    if province: community_reports_q = community_reports_q.filter(models.CrowdsourcedReport.province == province)
    community_reports_count = community_reports_q.scalar() or 0

    # 1. New articles count (total & needs verify)
    # [OPTIMIZATION] Combined two queries into one
    art_stats_q = db.query(
        func.count(Article.id),
        func.sum(case((Article.needs_verification.is_(True), 1), else_=0))
    ).filter(
        Article.published_at >= start,
        Article.published_at < end,
        Article.status == "approved"
    )
    if type: art_stats_q = art_stats_q.filter(Article.disaster_type == type)
    if province: art_stats_q = art_stats_q.filter(Article.province == province)
    if q: art_stats_q = art_stats_q.filter(Article.title.ilike(f"%{q}%"))
    
    art_stats_res = art_stats_q.first()
    total_articles = art_stats_res[0] or 0
    needs_verification_count = art_stats_res[1] or 0

    # 2. Events Aggregation (With Decision 18 Filtering)

    
    # [OPTIMIZATION] Reused centralized filter definition
    dec18_filter = get_dec18_filter()

    # Calculate Aggregates using SQL
    
    # For counts with conditions, we use case
    human_damage_case = case(
        ( (func.coalesce(Event.deaths, 0) + func.coalesce(Event.missing, 0) + func.coalesce(Event.injured, 0)) > 0, 1),
        else_=0
    )
    
    # Rough property damage check (billions > 0)
    prop_damage_case = case(
        ( func.coalesce(Event.damage_billion_vnd, 0) > 0, 1 ),
        else_=0
    )

    columns = [
        func.count(Event.id),
        func.sum(func.coalesce(Event.deaths, 0)),
        func.sum(func.coalesce(Event.missing, 0)),
        func.sum(func.coalesce(Event.injured, 0)),
        func.sum(human_damage_case),
        func.sum(prop_damage_case),
        func.sum(func.coalesce(Event.damage_billion_vnd, 0))
    ]
    
    # Reuse the same filter base logic for aggregation
    agg_q = db.query(*columns).filter(
        Event.last_updated_at >= start, 
        Event.started_at < end,
        Event.disaster_type != "unknown",
        Event.sources_count > 0,
        dec18_filter
    )
    if type: agg_q = agg_q.filter(Event.disaster_type == type)
    if province: agg_q = agg_q.filter(Event.province == province)
    if q: 
        q_clean = strip_accents(q).lower()
        agg_q = agg_q.filter(or_(
            Event.title.ilike(f"%{q}%"),
            Event.title.ilike(f"%{q_clean}%") # Fallback for unaccented search
        ))
    agg_q = get_visibility_filter(agg_q, is_admin)
        
    agg_res = agg_q.first()
    
    events_count = agg_res[0] or 0
    total_deaths = agg_res[1] or 0
    total_missing = agg_res[2] or 0
    total_injured = agg_res[3] or 0
    events_human_damage = agg_res[4] or 0
    events_property_damage = agg_res[5] or 0
    total_damage = agg_res[6] or 0

    # Type breakdown
    type_counts_q = db.query(Event.disaster_type, func.count(Event.id)).filter(
        Event.started_at >= start,
        Event.started_at < end,
        Event.disaster_type.notin_(["unknown"]),
        Event.sources_count > 0
    )
    if type: type_counts_q = type_counts_q.filter(Event.disaster_type == type)
    if province: type_counts_q = type_counts_q.filter(Event.province == province)
    if q: 
        q_clean = strip_accents(q).lower()
        type_counts_q = type_counts_q.filter(or_(
            Event.title.ilike(f"%{q}%"),
            Event.title.ilike(f"%{q_clean}%")
        ))
    type_counts_q = get_visibility_filter(type_counts_q, is_admin)
    
    type_counts_rows = type_counts_q.group_by(Event.disaster_type).all()
    
    official_types = list(DISASTER_GROUPS.keys())
    type_counts = {t: 0 for t in official_types}
    type_counts["community"] = 0
    type_counts["unknown"] = 0
    
    for row in type_counts_rows:
        dtype, cnt = row
        if dtype in type_counts:
            type_counts[dtype] += cnt
        else:
            type_counts["unknown"] += cnt # Should be 0 since we filtered unknown

    # Top Provinces breakdown (for hotspots) (Limit to top 20)
    prov_counts_q = db.query(Event.province, func.count(Event.id)).filter(
        Event.started_at >= start,
        Event.started_at < end,
        Event.disaster_type.notin_(["unknown"]),
        Event.sources_count > 0,
        Event.province.in_(PROVINCES)
    )
    if type: prov_counts_q = prov_counts_q.filter(Event.disaster_type == type)
    if province: prov_counts_q = prov_counts_q.filter(Event.province == province)
    if q: prov_counts_q = prov_counts_q.filter(Event.title.ilike(f"%{q}%"))
    prov_counts_q = get_visibility_filter(prov_counts_q, is_admin)
    
    prov_counts_rows = prov_counts_q.group_by(Event.province).order_by(func.count(Event.id).desc()).limit(20).all()
    by_province = [{"province": row[0], "events": row[1]} for row in prov_counts_rows]
    
    # Accurate count of ALL affected provinces
    provinces_count_q = db.query(func.count(func.distinct(Event.province))).filter(
        Event.started_at >= start,
        Event.started_at < end,
        Event.disaster_type.notin_(["unknown"]),
        Event.sources_count > 0,
        Event.province.in_(PROVINCES)
    )
    if type: provinces_count_q = provinces_count_q.filter(Event.disaster_type == type)
    if province: provinces_count_q = provinces_count_q.filter(Event.province == province)
    if q: 
        q_clean = strip_accents(q).lower()
        provinces_count_q = provinces_count_q.filter(or_(
            Event.title.ilike(f"%{q}%"),
            Event.title.ilike(f"%{q_clean}%")
        ))
    provinces_count_q = get_visibility_filter(provinces_count_q, is_admin)
    provinces_count = provinces_count_q.scalar() or 0

    res = {
        "window_hours": hours if not date else 24,
        "window_label": f"Ngày {date}" if date else f"Trong {hours}h qua",
        "articles_count": total_articles,
        "events_with_human_damage": int(events_human_damage),
        "events_with_property_damage": int(events_property_damage),
        "needs_verification_count": needs_verification_count,
        "events_count": events_count,
        "provinces_count": provinces_count,
        "community_reports_count": community_reports_count,
        "impacts": {
            "deaths": int(total_deaths),
            "missing": int(total_missing),
            "injured": int(total_injured),
            "damage_billion_vnd": float(total_damage)
        },
        "by_type": type_counts,
        "by_province": by_province,
    }
    cache.set(cache_key, res, ttl=120)

    # Cache summary for 2 mins (same as internal cache)
    if response: response.headers["Cache-Control"] = "public, max-age=120"
    return res

@router.get("/stats/timeline")
def stats_timeline(
    response: Response,
    hours: int = Query(24, ge=1, le=168), 
    date: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    type: str | None = Query(None),
    province: str | None = Query(None),
    q: str | None = Query(None),
    db: Session = Depends(get_db)
):
    """Timeline: số sự kiện theo giờ."""
    cache_key = f"timeline_{hours}_{date}_{start_date}_{end_date}_{type}_{province}_{q}"
    cached = cache.get(cache_key)
    if cached: 
        response.headers["Cache-Control"] = "public, max-age=180"
        return cached

    start, end = get_date_range(hours, date, start_date, end_date)
    
    # Database agnostic hour grouping
    if engine.url.drivername.startswith("postgresql"):
        time_func = func.date_trunc('hour', Event.started_at)
    else:
        time_func = func.strftime('%Y-%m-%d %H:00:00', Event.started_at)

    query = db.query(
        time_func.label('hour'),
        func.count(Event.id).label('count')
    ).filter(Event.started_at >= start, Event.started_at < end)
    
    if type: query = query.filter(Event.disaster_type == type)
    if province: query = query.filter(Event.province == province)
    if q:
        q_clean = strip_accents(q).lower()
        query = query.filter(or_(
            Event.title.ilike(f"%{q}%"),
            Event.title.ilike(f"%{q_clean}%")
        ))
    
    # [OPTIMIZATION] Sync visibility with dashboard
    query = get_visibility_filter(query, False) # Non-admin visibility for public timeline
    
    query = query.group_by('hour').order_by('hour')
    
    results = {row[0].strftime('%Y-%m-%d %H:00:00') if hasattr(row[0], 'strftime') else str(row[0]): row[1] for row in query.all()}
    
    # Fill gaps for the entire requested window for a continuous graph
    data = []
    current_dt = start
    while current_dt < end:
        h_str = current_dt.strftime('%Y-%m-%d %H:00:00')
        data.append({"time": h_str, "events": results.get(h_str, 0)})
        current_dt += timedelta(hours=1)
    
    res = {"window": date or f"{hours}h", "data": data}
    cache.set(cache_key, res, ttl=300)
    response.headers["Cache-Control"] = "public, max-age=300"
    return res


@router.get('/stream/events')
async def stream_events(request: Request):
    """Server-Sent Events endpoint streaming new events as they are published."""
    q = broadcast.subscribe()
    return StreamingResponse(broadcast.event_generator(q, request), media_type='text/event-stream')


# Admin logic moved to auth module consistent usage

@router.get('/admin/skip-logs')
def get_skip_logs(
    skip: int = Query(0, ge=0), 
    limit: int = Query(50, ge=1, le=200), 
    q: Optional[str] = Query(None),
    admin: models.User = Depends(auth.get_current_admin)
):
    """
    Fetch skipped/dropped articles from log file.
    Pagination supported via skip/limit.
    Items are returned in reverse chronological order (latest first).
    """
    logs_dir = Path(__file__).resolve().parents[1] / 'logs'
    log_file = logs_dir / 'review_potential_disasters.jsonl'
    
    if not log_file.exists():
        return []
        
    try:
        # Read all lines
        with log_file.open('r', encoding='utf-8') as f:
            lines = f.readlines()
            
        # Reverse to get latest first
        lines.reverse()
        
        # Filter if search query provided
        if q:
            q_lower = q.lower()
            filtered_lines = []
            for line in lines:
                try:
                    data = json.loads(line)
                    if q_lower in data.get("title", "").lower() or q_lower in data.get("url", "").lower():
                        filtered_lines.append(line)
                except:
                    continue
            lines = filtered_lines

        # Paginate
        start = skip
        end = skip + limit
        sliced_lines = lines[start:end]
        
        out = []
        for line in sliced_lines:
            try:
                out.append(json.loads(line))
            except Exception:
                continue
        return out
    except Exception as e:
        logger.error(f"Error reading skip logs: {e}")
        return []


@router.post('/admin/label')
def label_log(payload: dict, admin: models.User = Depends(auth.get_current_admin)):
    """Label a skipped/accepted item for training/audit. Payload must include `id` and `label`."""
    logs_dir = Path(__file__).resolve().parents[1] / 'logs'
    logs_dir.mkdir(parents=True, exist_ok=True)
    labels_file = logs_dir / 'labels.jsonl'
    record = {
        'timestamp': datetime.now(timezone.utc).isoformat(),
        'entry': payload,
    }
    try:
        with labels_file.open('a', encoding='utf-8') as f:
            f.write(json.dumps(record, ensure_ascii=False) + '\n')
    except Exception as e:
        return {'ok': False, 'error': str(e)}
    return {'ok': True}


@router.get('/admin/pending-articles')
def get_pending_articles(
    skip: int = 0, 
    limit: int = 50, 
    q: Optional[str] = Query(None),
    db: Session = Depends(get_db), 
    admin: models.User = Depends(auth.get_current_admin)
):
    query = db.query(models.Article).filter(models.Article.status == "pending")
    if q:
        query = query.filter(models.Article.title.ilike(f"%{q}%"))
    return query.order_by(models.Article.published_at.desc())\
                .offset(skip).limit(limit).all()

@router.get('/admin/rejected-articles')
def get_rejected_articles(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    q: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: models.User = Depends(auth.get_current_admin)
):
    """Fetch articles that were manually rejected."""
    query = db.query(models.Article).filter(models.Article.status == "rejected")
    if q:
        query = query.filter(models.Article.title.ilike(f"%{q}%"))
    
    return query.order_by(models.Article.moderated_at.desc())\
                .offset(skip).limit(limit).all()


@router.post('/admin/approve-article/{article_id}')
def approve_article(article_id: int, db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    """Approve a pending article and integrate it into events."""
    article = db.query(models.Article).filter(models.Article.id == article_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
        
    article.status = "approved"
    article.moderated_by = admin.id
    article.moderated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    
    # Update events
    ev, is_new = upsert_event_for_article(db, article)
    db.commit()
    
    log_audit("approve_article", admin.id, admin.email, event_id=ev.id, article_id=article_id)

    # Emit notifications post-commit
    emit_event_notifications(db, ev.id, is_new=is_new)
    # Invalidate event detail cache
    if article.event_id:
        cache.delete(f"ev_detail_v3_{article.event_id}_True")
        cache.delete(f"ev_detail_v3_{article.event_id}_False")
    
    # Global cache invalidation for consistency
    cache.delete_match("ev_v2_*")
    cache.delete_match("stats_*")
    cache.delete_match("articles_latest_*")
    cache.delete_match("map_*")
    cache.delete_match("heatmap_*")
        
    return {"ok": True, "message": "Article approved and event updated"}


@router.post('/admin/events/{event_id}/approve')
def approve_event(event_id: int, db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    """Approve an entire event: clear needs_verification and approve all articles."""
    ev = db.query(models.Event).filter(models.Event.id == event_id).with_for_update().first()
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
        
    ev.needs_verification = 0
    ev.confidence = 1.0 # Boost to ensure public visibility after manual review
    # Also approve all pending articles in this event
    db.query(models.Article).filter(
        models.Article.event_id == event_id,
        models.Article.status == "pending"
    ).update({"status": "approved"}, synchronize_session=False)
    
    db.commit()
    
    log_audit("approve_event", admin.id, admin.email, event_id=event_id)
    
    # [OPTIMIZATION] Real-time broadcast for Event Approval
    emit_event_notifications(db, event_id)

    # Invalidate cache
    cache.delete(f"ev_detail_v3_{event_id}_True")
    cache.delete(f"ev_detail_v3_{event_id}_False")
    cache.delete_match("ev_v2_*")
    cache.delete_match("stats_*")
    cache.delete_match("map_*")
    cache.delete_match("heatmap_*")
    
    return {"ok": True, "message": "Event and its articles approved"}


@router.get('/admin/blacklist')
def get_blacklist(
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db), 
    admin: models.User = Depends(auth.get_current_admin)
):
    """Fetch blacklisted article hashes."""
    return db.query(models.Blacklist).order_by(models.Blacklist.created_at.desc())\
             .offset(skip).limit(limit).all()

@router.delete('/admin/blacklist/{news_hash}')
def delete_from_blacklist(
    news_hash: str, 
    db: Session = Depends(get_db), 
    admin: models.User = Depends(auth.get_current_admin)
):
    """Remove a hash from the blacklist."""
    entry = db.query(models.Blacklist).filter(models.Blacklist.news_hash == news_hash).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Blacklist entry not found")
    db.delete(entry)
    db.commit()
    log_audit("delete_blacklist", admin.id, admin.email, details={"hash": news_hash})
    return {"ok": True}

@router.post('/admin/reject-article/{article_id}')
def reject_article(article_id: int, db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    """Reject an article and add its hash to blacklist to prevent re-crawling."""
    art = _reject_article_internal(db, article_id, admin, "Manual rejection via Admin UI")
    if not art:
        raise HTTPException(status_code=404, detail="Article not found")
    return {"ok": True, "message": "Article rejected and event updated"}

@router.post('/admin/unreject-article/{article_id}')
def unreject_article(article_id: int, db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    """Restore a rejected article, removing its hash from the blacklist and setting status back to pending."""
    art = db.query(Article).filter(Article.id == article_id).first()
    if not art:
        raise HTTPException(status_code=404, detail="Article not found")
        
    art.status = "pending"
    # Remove from blacklist if exists
    if art.news_hash:
        db.query(models.Blacklist).filter(models.Blacklist.news_hash == art.news_hash).delete()
    
    db.commit()
    log_audit("unreject_article", admin.id, admin.email, article_id=article_id)
    # Clear cache for pending list
    cache.delete_match("articles_latest_*")
    return {"ok": True, "message": "Article restored to pending status."}


@router.post('/alerts')
def post_alert(payload: dict, admin: models.User = Depends(auth.get_current_admin)):
    """Receive an alert to be pushed to subscribers or external push systems.
    This stores the alert in `logs/alerts.jsonl` and publishes it to SSE subscribers.
    """
    logs_dir = Path(__file__).resolve().parents[1] / 'logs'
    logs_dir.mkdir(parents=True, exist_ok=True)
    alerts_file = logs_dir / 'alerts.jsonl'
    record = {'timestamp': datetime.now(timezone.utc).isoformat(), 'alert': payload}
    try:
        with alerts_file.open('a', encoding='utf-8') as f:
            f.write(json.dumps(record, ensure_ascii=False) + '\n')
        # publish to subscribers
        try:
            broadcast.publish_event_sync({'type': 'alert', 'alert': payload})
        except Exception:
            pass
    except Exception as e:
        return {'ok': False, 'error': str(e)}
    try:
        from .ws import manager
        manager.broadcast_sync({"type": "ALERT", "data": payload})
    except Exception: pass
    
    return {'ok': True}

@router.get("/stats/heatmap")
def stats_heatmap(
    hours: int = Query(24, ge=1, le=168), 
    date: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(auth.get_current_user_optional)
):
    """Heatmap: số sự kiện theo tỉnh"""
    is_admin = current_user and current_user.role == "admin"
    cache_key = f"heatmap_{hours}_{date}_{start_date}_{end_date}_{is_admin}"
    cached = cache.get(cache_key)
    if cached: return cached

    start, end = get_date_range(hours, date, start_date, end_date)

    query = db.query(Event.province, func.count(Event.id))
    query = query.filter(Event.last_updated_at >= start, Event.started_at < end)
    
    query = get_visibility_filter(query, is_admin)
    query = apply_dashboard_filters(query, db)
    
    query = query.filter(Event.province.in_(PROVINCES))
    query = query.group_by(Event.province)
    
    result_rows = query.all()
    
    # Sort
    sorted_data = sorted([{"province": r[0], "events": r[1]} for r in result_rows], key=lambda x: x["events"], reverse=True)
    
    res = {"hours": hours, "data": sorted_data}
    cache.set(cache_key, res, ttl=300)
    return res

@router.get("/stats/top-risky-province")
def top_risky_province(
    hours: int = Query(24, ge=1, le=168),
    date: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(auth.get_current_user_optional)
):
    """Tỉnh nguy hiểm nhất"""
    is_admin = current_user and current_user.role == "admin"
    start, end = get_date_range(hours, date, start_date, end_date)
    
    query = db.query(
        Event.province,
        func.count(Event.id).label('count'),
        func.max(Event.last_updated_at).label('latest')
    ).filter(
        Event.started_at >= start,
        Event.started_at < end,
        Event.province != 'unknown'
    )
    
    query = get_visibility_filter(query, is_admin)
    query = apply_dashboard_filters(query, db)
    
    query = query.group_by(Event.province).order_by(desc(text('count')), desc(text('latest'))).limit(1)
    
    result = query.first()
    if result:
        return {
            "province": result[0],
            "events_24h": result[1],
            "latest_update": result[2]
        }
    return {"province": "Chưa có", "events_24h": 0, "latest_update": None}
@router.get("/stats/sources-health")
def sources_health():
    """Returns the latest report from the SourceMonitor."""
    logs_dir = Path(__file__).resolve().parents[1] / 'logs'
    report_file = logs_dir / 'source_status.json'
    
    # Look for logs in both possible locations (dev and docker)
    try:
        if not report_file.exists():
            backend_dir = Path(__file__).resolve().parents[1]
            report_file = backend_dir / 'logs' / 'source_status.json'
            
        if not report_file.exists():
            return {"error": "Report not generated yet."}
            
        with report_file.open('r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        return {"error": f"Failed to read report: {str(e)}"}

@router.get("/admin/crawler-status")
def get_crawler_status(db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    return db.query(CrawlerStatus).all()

@router.post("/admin/crawler/run")
async def trigger_crawler(db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    """Manually triggers a full crawl job, ensuring it doesn't collide with scheduled runs."""
    from .crawler import process_once
    
    # 1. Check lock immediately
    if not cache.acquire_lock("crawl_job_full", timeout=2700):
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS, 
            detail="Máy chủ đang thực hiện thu thập dữ liệu (theo lịch hoặc thủ công). Vui lòng đợi đến khi phiên hiện tại kết thúc."
        )
    
    # 2. Release lock immediately because `process_once` will re-acquire it properly
    cache.release_lock("crawl_job_full")
    
    # 3. Trigger in thread to bypass event loop collision with asyncio.run
    async def run_wrapper():
        try:
            await asyncio.to_thread(process_once)
        except Exception as e:
            logger.error(f"[MANUAL_CRAWL] Failed: {e}")

    asyncio.create_task(run_wrapper())
    return {"ok": True, "message": "Crawler started in background with lock-safety."}

@router.post("/admin/system/clear-cache")
def clear_system_cache(admin: models.User = Depends(auth.get_current_admin)):
    """Clears all cached API responses."""
    from .cache import cache
    if hasattr(cache, 'clear'):
        cache.clear()
    return {"ok": True, "message": "Cache cleared."}

@router.post("/admin/system/restart")
def restart_system(admin: models.User = Depends(auth.get_current_admin)):
    """Triggers a backend restart by touching a source file (requires --reload)."""
    main_file = Path(__file__).resolve().parent / "main.py"
    if main_file.exists():
        # Append a newline to trigger uvicorn reload
        with open(main_file, "a") as f:
            f.write("\n")
        log_audit("restart_backend", admin.id, admin.email)
        return {"ok": True, "message": "Backend restart triggered (reloader)."}
    return {"ok": False, "message": "Main file not found, restart failed."}



@router.post("/admin/ai-feedback")
def submit_ai_feedback(payload: dict, db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    article_id = payload.get("article_id")
    corrected_type = payload.get("corrected_type")
    
    article = db.query(Article).filter(Article.id == article_id).first()
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
        
    feedback = AiFeedback(
        article_id=article_id,
        user_id=admin.id,
        original_type=article.disaster_type,
        corrected_type=corrected_type,
        comment=payload.get("comment")
    )
    db.add(feedback)
    
    # Actually update the article as well (manual override)
    old_type = article.disaster_type
    article.disaster_type = corrected_type
    
    # [LOGIC] If this article belongs to an event, we might want to refresh the event's type
    if article.event_id:
        ev = db.query(Event).filter(Event.id == article.event_id).with_for_update().first()
        if ev:
            from .nlp import DISASTER_PRIORITY_MAP
            # Simple rule: if new type has higher priority than current event type, or current event type was the old article type
            if DISASTER_PRIORITY_MAP.get(corrected_type, 99) < DISASTER_PRIORITY_MAP.get(ev.disaster_type, 99) or ev.disaster_type == old_type:
                ev.disaster_type = corrected_type
            
            cache.delete_match(f"ev_detail_v3_{ev.id}*")
            emit_event_notifications(db, ev.id, is_new=False)

    db.commit()
    log_audit("submit_ai_feedback", admin.id, admin.email, article_id=article_id, details={"from": old_type, "to": corrected_type})
    
    return {"ok": True, "message": "Feedback saved and classification updated."}

@router.get("/admin/export/event/{event_id}")
def export_event_data(event_id: int, format: str = "excel", db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    import pandas as pd
    ev = db.query(Event).filter(Event.id == event_id).first()
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")
        
    articles = db.query(Article).filter(Article.event_id == event_id).all()
    
    # Disaster type mapping for Vietnamese names (used globally)
    # TYPE_MAP is now at module level
    
    data = []
    for art in articles:
        # Construct summary damage description
        damage_desc = []
        if art.deaths or art.missing or art.injured:
            damage_desc.append(f"{art.deaths or 0} người chết, {art.missing or 0} mất tích, {art.injured or 0} bị thương.")
        if art.damage_billion_vnd:
            damage_desc.append(f"Thiệt hại khoảng {art.damage_billion_vnd} tỷ VNĐ.")
        if art.summary:
            damage_desc.append(art.summary)
        
        # [FIX] Convert to VN time for display consistency
        vn_tz = timezone(timedelta(hours=7))
        pub_vn = art.published_at.replace(tzinfo=timezone.utc).astimezone(vn_tz)
        evt_vn = (art.event_time or art.published_at).replace(tzinfo=timezone.utc).astimezone(vn_tz)

        row = {
            "Loại hình thiên tai": TYPE_MAP.get(art.disaster_type, art.disaster_type),
            "Thời gian": evt_vn.strftime("%d/%m/%Y"),
            "Ngày đăng tin": pub_vn.strftime("%d/%m/%Y"),
            "Tuyến đường": art.route or "",
            "Vị trí thôn/bản": art.village or "",
            "Xã": art.commune or "",
            "Tỉnh": art.province or "",
            "Nguyên nhân (mưa hay hoạt động nhân sinh)": art.cause or "",
            "Mô tả đặc điểm trượt lở": art.characteristics or "",
            "Mô tả thiệt hại": " ".join(damage_desc),
            "Hình ảnh": art.image_url or "",
            "Nguồn": art.url
        }
        data.append(row)
        
    df = pd.DataFrame(data)
    
    if format == "excel":
        output = io.BytesIO()
        sheet_name = 'Chi tiết thiệt hại'
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            # [OPTIMIZATION] Auto-adjust column widths for Vietnamese content
            worksheet = writer.sheets[sheet_name]
            from openpyxl.utils import get_column_letter
            for i, column in enumerate(df.columns):
                # Calculate max length of data in column (safe for empty df)
                col_data_len = df[column].astype(str).map(len).max() if not df[column].empty else 0
                col_header_len = len(str(column))
                # Set width (header vs content, min 12, max 60 to prevent super wide cols)
                adjusted_width = min(max(col_data_len, col_header_len) + 3, 60)
                worksheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_width
        
        output.seek(0)
        
        headers = {'Content-Disposition': f'attachment; filename="bao-cao-thiet-hai-event-{event_id}.xlsx"'}
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    elif format == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # Try to use a font that supports Vietnamese if available, otherwise fallback
        # To avoid squares/error, we strip accents for PDF export as a safer fallback
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        clean_title = strip_accents(ev.title)
        elements.append(Paragraph(f"BAO CAO THIET HAI SU KIEN: {clean_title}", styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Thoi gian bat dau: {ev.started_at.strftime('%Y-%m-%d')}", styles['Normal']))
        elements.append(Paragraph(f"Loai thien tai: {strip_accents(ev.disaster_type)} | Tinh thanh: {strip_accents(ev.province)}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table data
        table_data = [["Ngay", "Nguon", "Tieu de", "Tu vong", "Mat tich", "Bi thuong", "Thiet hai"]]
        for art in articles:
            table_data.append([
                art.published_at.strftime("%d/%m"),
                strip_accents(art.source)[:15],
                strip_accents(art.title)[:50] + "...",
                str(art.deaths or 0),
                str(art.missing or 0),
                str(art.injured or 0),
                str(art.damage_billion_vnd or 0)
            ])
            
        t = Table(table_data, colWidths=[50, 80, 400, 50, 50, 50, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        headers = {'Content-Disposition': f'attachment; filename="bao-cao-thiet-hai-event-{event_id}.pdf"'}
        return StreamingResponse(buffer, headers=headers, media_type='application/pdf')

@router.get("/admin/export/daily")
def export_daily_summary(date: str = None, db: Session = Depends(get_db), admin: models.User = Depends(auth.get_current_admin)):
    import pandas as pd
    import calendar
    
    vn_tz = timezone(timedelta(hours=7))

    # [FIX] Timezone-aware date parsing for Vietnam (UTC+7)
    # Ensure "Daily" covers 00:00-23:59 VN time, not UTC.
    now_utc = datetime.now(timezone.utc)
    now_vn = now_utc.astimezone(vn_tz)
    
    if not date:
        target_date_vn = now_vn.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        # Parse YYYY-MM-DD as VN midnight
        d = datetime.strptime(date, "%Y-%m-%d")
        target_date_vn = d.replace(tzinfo=vn_tz)
        
    start_utc = target_date_vn.astimezone(timezone.utc)
    end_utc = start_utc + timedelta(days=1)
    
    events = db.query(Event).filter(Event.started_at >= start_utc, Event.started_at < end_utc).options(selectinload(Event.articles)).all()
    
    data = []
    for ev in events:
        # Get representative URL
        ref_link = ev.articles[0].url if ev.articles else ""

        data.append({
            "Ngày": ev.started_at.astimezone(vn_tz).strftime("%d/%m/%Y"),
            "Loại hình": TYPE_MAP.get(ev.disaster_type, ev.disaster_type),
            "Tỉnh thành": ev.province,
            "Xã/Thôn/Tuyến": f"{ev.commune or ''} / {ev.village or ''} / {ev.route or ''}".strip(" /"),
            "Nguyên nhân": ev.cause or "Chưa rõ",
            "Đặc điểm": ev.characteristics or "N/A",
            "Tử vong": ev.deaths or 0,
            "Mất tích": ev.missing or 0,
            "Bị thương": ev.injured or 0,
            "Thiệt hại (Tỷ VNĐ)": ev.damage_billion_vnd or 0,
            "Link nguồn": ref_link
        })
        
    df = pd.DataFrame(data)
    output = io.BytesIO()
    display_date = target_date_vn.strftime("%d-%m-%Y")
    sheet_name = f'Báo-cáo-{display_date}'
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        # [OPTIMIZATION] Auto-adjust column widths
        worksheet = writer.sheets[sheet_name]
        from openpyxl.utils import get_column_letter
        for i, column in enumerate(df.columns):
            col_data_len = df[column].astype(str).map(len).max() if not df[column].empty else 0
            col_header_len = len(str(column))
            adjusted_width = min(max(col_data_len, col_header_len) + 3, 60)
            worksheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_width

    output.seek(0)
    
    headers = {'Content-Disposition': f'attachment; filename="bao-cao-ngay-{display_date}.xlsx"'}
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@router.get("/admin/export/summary")
def export_events_summary(
    start_date: str = Query(None),
    end_date: str = Query(None),
    month: int = Query(None),
    year: int = Query(None),
    type: str = Query(None),
    province: str = Query(None),
    q: str = Query(None),
    db: Session = Depends(get_db),
    admin: models.User = Depends(auth.get_current_admin)
):
    import pandas as pd
    import calendar
    
    vn_tz = timezone(timedelta(hours=7))

    # TYPE_MAP is used from module level

    query = db.query(Event).filter(Event.sources_count > 0)
    
    # Range Logic [FIXED TIMEZONE]
    title_suffix = ""
    if month and year:
        # Month start: 00:00 VN on 1st
        start_vn = datetime(year, month, 1, tzinfo=vn_tz)
        # Month end: 23:59:59 VN on last day
        last_day = calendar.monthrange(year, month)[1]
        end_vn = datetime(year, month, last_day, 23, 59, 59, tzinfo=vn_tz)
        
        start_utc = start_vn.astimezone(timezone.utc)
        end_utc = end_vn.astimezone(timezone.utc)
        
        query = query.filter(Event.started_at >= start_utc, Event.started_at <= end_utc)
        title_suffix = f"Tháng {month}-{year}"
    elif start_date:
        # Start date: 00:00 VN
        sd_vn = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=vn_tz)
        start_utc = sd_vn.astimezone(timezone.utc)
        
        if end_date:
            # End date: 23:59:59 VN (so next day 00:00)
            ed_vn = (datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)).replace(tzinfo=vn_tz)
            end_utc = ed_vn.astimezone(timezone.utc)
            query = query.filter(Event.started_at >= start_utc, Event.started_at < end_utc)
            title_suffix = f"Từ {start_date} Đến {end_date}"
        else:
            # Single day
            ed_utc = start_utc + timedelta(days=1)
            query = query.filter(Event.started_at >= start_utc, Event.started_at < ed_utc)
            title_suffix = f"Ngày {start_date}"

    if type: query = query.filter(Event.disaster_type == type)
    if province: query = query.filter(Event.province == province)
    if q:
        q_clean = strip_accents(q).lower()
        query = query.filter(or_(
            Event.title.ilike(f"%{q}%"),
            Event.title.ilike(f"%{q_clean}%")
        ))

    events = query.order_by(Event.started_at.desc()).options(selectinload(Event.articles)).all()
    
    data = []
    for ev in events:
        ref_link = ev.articles[0].url if ev.articles else ""
        data.append({
            "Ngày Báo cáo": ev.started_at.astimezone(vn_tz).strftime("%d/%m/%Y"),
            "Tên sự kiện": ev.title,
            "Loại hình thiên tai": TYPE_MAP.get(ev.disaster_type, ev.disaster_type),
            "Tỉnh": ev.province,
            "Xã/Thôn/Tuyến": f"{ev.commune or ''} / {ev.village or ''} / {ev.route or ''}".strip(" /"),
            "Nguyên nhân": ev.cause or "Chưa rõ",
            "Tử vong": ev.deaths or 0,
            "Mất tích": ev.missing or 0,
            "Bị thương": ev.injured or 0,
            "Thiệt hại (Tỷ VNĐ)": ev.damage_billion_vnd or 0,
            "Link nguồn": ref_link,
            "Trạng thái": "Đã duyệt" if ev.needs_verification == 0 else "Chờ xác minh"
        })
        
    df = pd.DataFrame(data)
    
    if not df.empty:
        summary_row = {
            "Ngày Báo cáo": "TỔNG CỘNG",
            "Tên sự kiện": title_suffix,
            "Loại hình thiên tai": "",
            "Tỉnh": "",
            "Xã/Thôn/Tuyến": "",
            "Nguyên nhân": "",
            "Tử vong": df["Tử vong"].sum(),
            "Mất tích": df["Mất tích"].sum(),
            "Bị thương": df["Bị thương"].sum(),
            "Thiệt hại (Tỷ VNĐ)": df["Thiệt hại (Tỷ VNĐ)"].sum(),
            "Link nguồn": "",
            "Trạng thái": ""
        }
        df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)

    output = io.BytesIO()
    sheet_name = 'Tổng hợp'
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        # [OPTIMIZATION] Auto-adjust column widths
        worksheet = writer.sheets[sheet_name]
        from openpyxl.utils import get_column_letter
        for i, column in enumerate(df.columns):
            col_data_len = df[column].astype(str).map(len).max() if not df[column].empty else 0
            col_header_len = len(str(column))
            adjusted_width = min(max(col_data_len, col_header_len) + 3, 60)
            worksheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_width

    output.seek(0)
    
    filename = f"bao-cao-thiet-hai-{title_suffix.replace(' ', '_')}.xlsx"
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
